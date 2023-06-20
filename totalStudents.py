import openpyxl
import sys
import json

data_json = sys.argv[1]
data = json.loads(data_json)

details = data['details']
sem = details[0].get("sem")
s = set()
for i in range(len(details)):
    s.add(details[i].get("slot"))
slot_list = list(s)
noOfStudents = 0
for i in slot_list:
    input_slot = i
    p = 1
    b = 1
    code_list = list()
    for i in range(len(details)):
        if details[i].get("slot") == input_slot:
            code_list.append(details[i].get("branch"))

    # print(code_list)
    for code in code_list:
        check_supply = 1
        wb_branch = openpyxl.load_workbook('.\\updatedExcels\\S'+str(sem)+'_'+code+'.xlsx')
        ws_branch_reg = wb_branch[input_slot]
        # CHECKING IF THERE IS SUPPLY STUDENTS OR NOT
        try:
            ws_branch_sply = wb_branch[input_slot+'_supply']
        except:
            check_supply = 0
        # COUNT OF NORMAL STUDENTS
        for r in range(1, ws_branch_reg.max_row+1):
            noOfStudents += 1
        if (check_supply == 1):
            # COUNT OF SUPPLY STUDENTS
            for r in range(1, ws_branch_sply.max_row+1):
                noOfStudents += 1
rem = noOfStudents % 30
noOfRooms30 = int(noOfStudents - rem)//30
noOfRooms60 = 0
if rem <= 10:
    noOfRooms30 -= 1
    noOfRooms60 = 1
elif rem >= 20:
    noOfRooms30 += 1
else:
    noOfRooms30 -= 2
    noOfRooms60 = 2
# NO: OF STUDENTS WRITING EXAM FOR A PARTICULAR SLOT
print(noOfStudents)
# print(noOfRooms30)
# print(noOfRooms60)
