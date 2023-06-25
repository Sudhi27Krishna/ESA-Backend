import openpyxl
import math
from openpyxl import Workbook
import sys
import json

# ACCEPT JSOM DATA FROM COMMAND-LINE ARGUMENTS
data_json = sys.argv[1]
data = json.loads(data_json)

for sem in data:
    # LOADING MAIN EXCEL SHEET
    # wb = openpyxl.load_workbook('./uploadedExcels/arjun-miniproject.xlsx')
    wb = openpyxl.load_workbook('.\\uploadedExcels\\'+sem)
    sheetname = wb.sheetnames
    ws = wb[sheetname[0]]

    # IDENTIFYING BRANCHES
    branches = set()
    for a in ws['D']:
        branches.add(a.value)
    # print(branches)
    branch_list = list(branches)
    while None in branch_list:
        # removing None from list using remove method
        branch_list.remove(None)
    if 'Branch Name' in branch_list:
        branch_list.remove('Branch Name')
    print(branch_list)
    code_list = list()
    sub_list = list()
    # CREATING WORKBOOKS FOR EACH BRANCH
    for sub in branch_list:
        wb_branch = Workbook()
        ws_branchMain = wb_branch.active
        ws_branchMain.title = 'Main'
        r = 1
        # print(ws.max_row)
        print('\n', sub)
        for p in range(1, ws.max_row+1):
            if (ws.cell(row=p, column=4).value == sub):
                nm = ws.cell(row=p, column=1).value
                regno = nm[-11:-1:1]
                ws_branchMain.cell(row=r, column=1).value = nm
                ws_branchMain.cell(row=r, column=2).value = regno
                ws_branchMain.cell(row=r, column=3).value = ws.cell(
                    row=p, column=4).value
                ws_branchMain.cell(row=r, column=4).value = ws.cell(
                    row=p, column=7).value
                ws_branchMain.cell(row=r, column=5).value = ws.cell(
                    row=p, column=8).value[-9:-3:1]
                r += 1
        codeno = regno[5:7]
        # print(ws_branchMain.max_row,codeno)
        code_list.append(codeno)
        xl = '.\\updatedExcels\\'+sem[0:2]+'_'+codeno+'.xlsx'
        wb_branch.save(xl)
        # sorting in alphabetical order
        wb_branch.create_sheet('Temp')
        wb_branchMain = wb_branch['Temp']
        s = set()
        for r in ws_branchMain.iter_rows(min_row=1, values_only=True):
            s.add(r)
        li = list(s)
        li1 = sorted(li, key=lambda x: x[1])
        for row in li1:
            wb_branchMain.append(row)

        # IDENTIFYING BRANCHES
        slot_set = set()
        for a in wb_branchMain['D']:
            slot_set.add(a.value)
        # print(slot_set)
        slot_list = list(slot_set)
        print(slot_list)
        slot_list.sort()
        thisdict = dict()
        # CREATING SHEETS FOR EACH SLOT IN A BRANCH
        for slot in slot_list:
            wb_branch.create_sheet(slot)
            wb_branchRegular = wb_branch[slot]
            year_set = set()
            print(slot)
            for p in range(1, wb_branchMain.max_row+1):
                if (wb_branchMain.cell(row=p, column=4).value == slot):
                    year_set.add(wb_branchMain.cell(
                        row=p, column=2).value[3:5])

            year_list = list(year_set)
            year_list.sort()
            # print('Total years: ',year_set,year_list)
            if len(year_list) != 1:
                wb_branch.create_sheet(slot+'_supply')
                wb_branchSupply = wb_branch[slot+'_supply']
            r = 1
            check_supply = 1
            for p in range(1, wb_branchMain.max_row+1):
                if (wb_branchMain.cell(row=p, column=4).value == slot):
                    if (wb_branchMain.cell(row=p, column=2).value[3:5] == year_list[-1]):
                        wb_branchRegular.cell(row=r, column=1).value = wb_branchMain.cell(
                            row=p, column=1).value
                        wb_branchRegular.cell(row=r, column=2).value = wb_branchMain.cell(
                            row=p, column=2).value
                        wb_branchRegular.cell(row=r, column=3).value = wb_branchMain.cell(
                            row=p, column=4).value
                        wb_branchRegular.cell(row=r, column=3).value = wb_branchMain.cell(
                            row=p, column=5).value
                        r += 1
                    else:
                        wb_branchSupply.cell(row=check_supply, column=1).value = wb_branchMain.cell(
                            row=p, column=1).value
                        wb_branchSupply.cell(row=check_supply, column=2).value = wb_branchMain.cell(
                            row=p, column=2).value
                        wb_branchSupply.cell(row=check_supply, column=3).value = wb_branchMain.cell(
                            row=p, column=4).value
                        wb_branchSupply.cell(row=check_supply, column=3).value = wb_branchMain.cell(
                            row=p, column=5).value
                        check_supply += 1
            print('Max row in Regular and supply: ', wb_branchRegular.max_row)
            thisdict[codeno+slot] = wb_branchRegular.max_row
            if check_supply > 1:
                print(wb_branchSupply.max_row)
        wb_branch.save(xl)
        sub_list.append(thisdict)
