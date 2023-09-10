import openpyxl
import math
from openpyxl import Workbook
from openpyxl.styles import Alignment 
import sys
import json

# ACCEPT JSOM DATA FROM COMMAND-LINE ARGUMENTS
data_json = sys.argv[1]
data = json.loads(data_json)

# TAKING DATAS FROM DATA DICTIONARY
rooms = list()
details = list()
date = data['date']
time = data['time']
rooms = data['rooms']
details = data['details']
slot_set = set()
for i in range(len(details)):
    slot_set.add(details[i].get("slot"))
slot_list = list(slot_set)

print(rooms)
no60=0;no30=0
for i in range(len(rooms)):
    if(rooms[i].get("capacity")==60):
        no60+=1
no30=len(rooms)-no60
print(no60)

for i in slot_list:
    input_slot = i
    # CREATING WORKBOOK FOR GIVEN SLOT AND BRANCHES
    wb_slot = Workbook()
    ws_slotMain = wb_slot.active
    ws_slotMain.title = input_slot
    wb_slot.save('.\\updatedExcels\\'+date+'_'+time+'.xlsx')
    wb_slot.create_sheet('balance')
    ws_slot_splyMain = wb_slot['balance']
    p = 1
    b = 1
    code_list = list()
    for i in range(len(details)):
        if (details[i].get("slot") == input_slot):
            code_list.append(details[i].get("branch"))
    print(code_list)
    s = details[0].get("sem")
    for code in code_list:
        check_supply = 1
        wb_branch = openpyxl.load_workbook(
            '.\\updatedExcels\\'+'S'+str(s)+'_'+code+'.xlsx')
        ws_branch_reg = wb_branch[input_slot]
        # CHECKING IF THERE IS SUPPLY STUDENTS OR NOT
        try:
            ws_branch_sply = wb_branch[input_slot+'_supply']
        except:
            check_supply = 0
        check = ws_branch_reg.max_row-(ws_branch_reg.max_row % 30)
        # NO: OF REGULAR STUDENTS THAT CAN BE PLACED IN ROOMS WHICH FITS EXACTLY 15 STUDENTS FROM SINGLE BRANCH i.e slot SHEET
        for r in range(1, check+1):
            # print(ws_branch_reg.max_row-r, ws_branch_reg.max_row, r, p, b, check)
            ws_slotMain.cell(row=p, column=1).value = ws_branch_reg.cell(
                row=r, column=3).value
            ws_slotMain.cell(row=p, column=2).value = ws_branch_reg.cell(
                row=r, column=2).value
            p += 1
        # NO: OF REGULAR STUDENTS WHO WILL NOT FIT IN THE ABOVE CASE SO ADDING THEM WITH SUPPLY STUDENTS  i.e BALANCE SHEET
        for r in range(check+1, ws_branch_reg.max_row+1):
            ws_slot_splyMain.cell(row=b, column=1).value = ws_branch_reg.cell(
                row=r, column=3).value
            ws_slot_splyMain.cell(row=b, column=2).value = ws_branch_reg.cell(
                row=r, column=2).value
            b += 1
        # IF THERE IS SUPPLY STUDENTS ADDING THEM TO BALANCE SHEET
        if (check_supply == 1):
            for r in range(1, ws_branch_sply.max_row+1):
                ws_slot_splyMain.cell(row=b, column=1).value = ws_branch_sply.cell(
                    row=r, column=3).value
                ws_slot_splyMain.cell(row=b, column=2).value = ws_branch_sply.cell(
                    row=r, column=2).value
                b += 1
        wb_slot.save('.\\updatedExcels\\'+date+'_'+time+'.xlsx')

    # normal arrangement
    # for i in range(1, math.ceil(ws_slotMain.max_row/30)+1):
    for i in range(1, no30+1):
        room_sheet = wb_slot.create_sheet('rno'+str(i))
    wb_slot.save('.\\updatedExcels\\'+date+'_'+time+'.xlsx')
    sh_nm = wb_slot.sheetnames
    del sh_nm[0]
    del sh_nm[0]
    max_sheets = len(sh_nm)
    no30rem=math.ceil(ws_slotMain.max_row/30)-no30
    print(no30rem)
    # print(ws_slot_splyMain.max_row)
    if no30rem>0:
        b=ws_slot_splyMain.max_row+1
        # print(ws_slotMain.cell(row=no30*30+1, column=1).value)
        for i in range(no30*30+1,ws_slotMain.max_row+1):
            ws_slot_splyMain.cell(row=b, column=1).value = ws_slotMain.cell(
                row=i, column=1).value
            ws_slot_splyMain.cell(row=b, column=2).value = ws_slotMain.cell(
                row=i, column=2).value
            b+=1
        ws_slotMain.delete_rows(no30*30+1, ws_slotMain.max_row-1)
    j = 0
    p = 1
    ch = 0
    # cnt = 0
    # print(ws_slotMain.max_row)
    for i in range(1, ws_slotMain.max_row+1, 15):
        # print(i,cnt)
        room_sheet = wb_slot[sh_nm[j]]
        # print(sh_nm[j]+"  "+str(room_sheet.max_row))
        for r in range(1, 16):
            room_sheet.cell(row=p, column=1).value = ws_slotMain.cell(
                row=(ch*15+r), column=1).value
            room_sheet.cell(row=p, column=2).value = ws_slotMain.cell(
                row=(ch*15+r), column=2).value
            p += 1
            # cnt+=1
        ch += 1
        if (ch >= max_sheets):
            p = 16
        else:
            p = 1
        # print(sh_nm[j]+"  "+str(room_sheet.max_row))
        j = (ch) % max_sheets
        # wb_slot.save('.\\updatedExcels\\'+date+'_'+time+'.xlsx')
    wb_slot.save('.\\updatedExcels\\'+date+'_'+time+'.xlsx')
    # print(cnt)
    ch = 0
    # balance students to normal class
    bal = ws_slot_splyMain.max_row
    rem = bal % 15
    for j in range(max_sheets-1, 1, -1):
        room_sheet = wb_slot[sh_nm[j]]
        if bal==rem:
            gen=rem
        else:
            gen=15
        if (room_sheet.max_row == 15):
            # print("_________"+sh_nm[j]+"  "+str(room_sheet.max_row))
            p = 16
            for r in range(1, gen):
                room_sheet.cell(row=p, column=1).value = ws_slot_splyMain.cell(
                    row=(ch*15+r), column=1).value
                room_sheet.cell(row=p, column=2).value = ws_slot_splyMain.cell(
                    row=(ch*15+r), column=2).value
                p += 1
        else:
            break
        bal-=15
        ch += 1
    wb_slot.save('.\\updatedExcels\\'+date+'_'+time+'.xlsx')

    # balance students to drawing room 
    bal = ws_slot_splyMain.max_row-ch*15
    ch = ch*15
    rem = bal % 30
    extra_rooms = no60
    for i in range(max_sheets+1, max_sheets+extra_rooms+1):
        room_sheet = wb_slot.create_sheet('rno'+str(i))
    wb_slot.save('.\\updatedExcels\\'+date+'_'+time+'.xlsx')
    i1 = ch
    j = -1
    p = 1
    ch = 0
    for i in range(i1+1, ws_slot_splyMain.max_row+1, 30):
        room_sheet = wb_slot[wb_slot.sheetnames[j]]
        if bal==rem:
            gen=rem
        else:
            gen=30
        print(gen,ch,i,i+gen,ch*30+i)
        for r in range(i,i+gen):
            room_sheet.cell(row=p, column=1).value = ws_slot_splyMain.cell(
                row=r, column=1).value
            room_sheet.cell(row=p, column=2).value = ws_slot_splyMain.cell(
                row=r, column=2).value
            p += 1
        bal-=30
        ch += 1
        if (ch >= extra_rooms):
            p = 31
        else:
            p = 1
        j = (ch) % extra_rooms
        j += 1
        j *= -1
        wb_slot.save('.\\updatedExcels\\'+date+'_'+time+'.xlsx')
    
    # RENAMING ROOMS
    sh_nm = wb_slot.sheetnames
    del sh_nm[0]
    del sh_nm[0]
    for j in sh_nm:
        room_sheet = wb_slot[j]
        room_sheet.insert_cols(idx=1)
        room_sheet.insert_cols(idx=1)
        max_alloc1 = room_sheet.max_row//2
        max_alloc2 = room_sheet.max_row-max_alloc1
        wb_slot.save('.\\updatedExcels\\'+date+'_'+time+'.xlsx')
        li = list()
        for i in range(1, max_alloc1+1):
            li.append('A'+str(i))
        for i in range(1, max_alloc2+1):
            li.append('B'+str(i))
        r = 1
        for i in li:
            room_sheet.cell(row=r, column=2).value = i
            r += 1
        for i in range(len(rooms)):
            if ((room_sheet.max_row <= rooms[i].get("capacity"))):
                room_sheet.title = rooms[i].get("room_no")
                # print(rooms[i].get("room_no"))
                del rooms[i]
                wb_slot.save('.\\updatedExcels\\'+date+'_'+time+'.xlsx')
                break
    wb_slot.save('.\\updatedExcels\\'+date+'_'+time+'.xlsx')

    #BRANCH DIVISION IN EXCEL
    sh_nm = wb_slot.sheetnames
    del sh_nm[0]
    del sh_nm[0]
    for j in sh_nm:
        room_sheet = wb_slot[j]
        i=0;start=2;flag=0
        room_sheet.insert_rows(idx=1)
        room_sheet.cell(row=1,column=1,value='Branch')
        room_sheet.cell(row=1,column=2,value='Seat No.')
        room_sheet.cell(row=1,column=3,value='Sub. code')
        room_sheet.cell(row=1,column=4,value='Reg. No')
        br=room_sheet.cell(row=2, column=4).value[5:7]
        for p in range(2,room_sheet.max_row+1):
                    # print(room_sheet.cell(row=p, column=4).value[5:7])
                    if(br!=room_sheet.cell(row=p, column=4).value[5:7] or p==room_sheet.max_row):
                        if(p!=room_sheet.max_row):
                            s='A'+str(start)+':A'+str(start+i-1)
                        else:
                            s='A'+str(start)+':A'+str(start+i)
                        room_sheet.merge_cells(s)
                        cell = room_sheet.cell(row=start,column=1)
                        cell.value = br  
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        wb_slot.save('.\\updatedExcels\\'+date+'_'+time+'.xlsx')
                        flag=1
                    if(flag==1):    
                        br=room_sheet.cell(row=p, column=4).value[5:7]
                        start=p
                        i=0
                        flag=0
                    i+=1
        wb_slot.save('.\\updatedExcels\\'+date+'_'+time+'.xlsx')
